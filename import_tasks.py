#!/usr/bin/env python3
"""
Import tasks from Excel file into AGL-Tracking format.
This script converts Excel tasks into PPM tasks compatible with the AGL MCT system.
"""

import openpyxl
import json
from datetime import datetime, timedelta
import random

def generate_task_id():
    """Generate a unique task ID based on timestamp."""
    return int(datetime.now().timestamp() * 1000) + random.randint(0, 999)

def parse_task_description(task_text):
    """
    Parse task description to extract meaningful information.
    Format: "PPM for <Equipment> at <Location>"
    """
    parts = task_text.replace("PPM for ", "").split(" at ")
    if len(parts) == 2:
        equipment = parts[0].strip()
        location = parts[1].strip()
        return equipment, location
    return task_text, ""

def determine_frequency(description):
    """Determine frequency based on task description."""
    desc_lower = description.lower()
    
    # Common patterns for different frequencies
    if any(word in desc_lower for word in ['daily', 'every day']):
        return 'Daily'
    elif any(word in desc_lower for word in ['weekly', 'every week']):
        return 'Weekly'
    elif any(word in desc_lower for word in ['monthly', 'every month']):
        return 'Monthly'
    elif any(word in desc_lower for word in ['quarterly', 'every quarter']):
        return 'Quarterly'
    elif any(word in desc_lower for word in ['yearly', 'annual', 'every year']):
        return 'Yearly'
    
    # Default frequency for maintenance tasks
    return 'Weekly'

def determine_task_type(description):
    """Determine task type based on description."""
    desc_lower = description.lower()
    
    if any(word in desc_lower for word in ['inspect', 'check', 'review', 'examine']):
        return 'Inspection'
    elif any(word in desc_lower for word in ['clean', 'wash', 'sanitize']):
        return 'Cleaning'
    elif any(word in desc_lower for word in ['repair', 'fix', 'replace']):
        return 'Repair'
    elif any(word in desc_lower for word in ['service', 'maintain', 'ppm']):
        return 'Service'
    elif any(word in desc_lower for word in ['test', 'calibrate', 'measure']):
        return 'Testing'
    
    # Default to Inspection for PPM tasks
    return 'Inspection'

def calculate_due_date(start_date, index, frequency):
    """Calculate due date based on frequency and index."""
    if frequency == 'Daily':
        return (start_date + timedelta(days=index)).strftime('%Y-%m-%d')
    elif frequency == 'Weekly':
        return (start_date + timedelta(weeks=index)).strftime('%Y-%m-%d')
    elif frequency == 'Monthly':
        # Approximate monthly as 30 days
        return (start_date + timedelta(days=30 * index)).strftime('%Y-%m-%d')
    elif frequency == 'Quarterly':
        return (start_date + timedelta(days=90 * index)).strftime('%Y-%m-%d')
    elif frequency == 'Yearly':
        return (start_date + timedelta(days=365 * index)).strftime('%Y-%m-%d')
    else:
        return (start_date + timedelta(days=7 * index)).strftime('%Y-%m-%d')

def import_tasks_from_excel(filename):
    """Import tasks from Excel and convert to AGL-Tracking format."""
    wb = openpyxl.load_workbook(filename)
    
    print("=" * 80)
    print("🔄 AGL-Tracking Task Import Utility")
    print("=" * 80)
    
    ppm_tasks = []
    start_date = datetime.now()
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n📄 Processing Sheet: {sheet_name}")
        
        max_row = ws.max_row
        task_count = 0
        
        # Skip header row, start from row 2
        for row_idx in range(2, max_row + 1):
            task_description = ws.cell(row_idx, 1).value
            
            if not task_description or task_description.strip() == "":
                continue
            
            # Parse task description
            equipment, location = parse_task_description(task_description)
            
            # Determine task attributes
            frequency = determine_frequency(task_description)
            task_type = determine_task_type(task_description)
            
            # Calculate staggered due dates to spread out tasks
            due_date = calculate_due_date(start_date, task_count % 30, frequency)
            
            # Create PPM task object matching app.js structure
            ppm_task = {
                "id": generate_task_id() + task_count,
                "description": task_description,
                "shiftType": "Both",  # Default to both shifts
                "type": task_type,
                "dueDate": due_date,
                "frequency": frequency,
                "status": "Not Started",
                "dayShift": "",  # To be assigned
                "nightShift": "",  # To be assigned
                "photos": [],
                "lastCompleted": None
            }
            
            ppm_tasks.append(ppm_task)
            task_count += 1
        
        print(f"   ✅ Imported {task_count} tasks from {sheet_name}")
    
    print(f"\n{'=' * 80}")
    print(f"📊 Import Summary")
    print(f"{'=' * 80}")
    print(f"Total Tasks Imported: {len(ppm_tasks)}")
    print(f"Start Date: {start_date.strftime('%Y-%m-%d')}")
    
    # Count by frequency
    freq_counts = {}
    for task in ppm_tasks:
        freq = task['frequency']
        freq_counts[freq] = freq_counts.get(freq, 0) + 1
    
    print(f"\nTasks by Frequency:")
    for freq, count in sorted(freq_counts.items()):
        print(f"   {freq}: {count} tasks")
    
    # Count by type
    type_counts = {}
    for task in ppm_tasks:
        task_type = task['type']
        type_counts[task_type] = type_counts.get(task_type, 0) + 1
    
    print(f"\nTasks by Type:")
    for task_type, count in sorted(type_counts.items()):
        print(f"   {task_type}: {count} tasks")
    
    return ppm_tasks

def export_to_json(tasks, output_file='imported_ppm_tasks.json'):
    """Export tasks to JSON format."""
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(tasks, f, indent=2, ensure_ascii=False)
    
    print(f"\n{'=' * 80}")
    print(f"✅ Tasks exported to: {output_file}")
    print(f"{'=' * 80}")

def generate_import_html(tasks):
    """Generate HTML file with JavaScript to import tasks into the app."""
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Import Tasks - AGL Tracking</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            max-width: 800px;
            margin: 50px auto;
            padding: 20px;
            background: #f5f5f5;
        }}
        .container {{
            background: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #2c3e50;
            border-bottom: 3px solid #3498db;
            padding-bottom: 10px;
        }}
        .info {{
            background: #e8f4f8;
            padding: 15px;
            border-left: 4px solid #3498db;
            margin: 20px 0;
        }}
        .warning {{
            background: #fff3cd;
            padding: 15px;
            border-left: 4px solid #ffc107;
            margin: 20px 0;
        }}
        .success {{
            background: #d4edda;
            padding: 15px;
            border-left: 4px solid #28a745;
            margin: 20px 0;
            display: none;
        }}
        button {{
            background: #3498db;
            color: white;
            padding: 12px 30px;
            border: none;
            border-radius: 5px;
            cursor: pointer;
            font-size: 16px;
            margin: 10px 5px;
        }}
        button:hover {{
            background: #2980b9;
        }}
        button.danger {{
            background: #e74c3c;
        }}
        button.danger:hover {{
            background: #c0392b;
        }}
        #progress {{
            display: none;
            margin: 20px 0;
        }}
        .progress-bar {{
            width: 100%;
            height: 30px;
            background: #ecf0f1;
            border-radius: 15px;
            overflow: hidden;
        }}
        .progress-fill {{
            height: 100%;
            background: #3498db;
            transition: width 0.3s;
            display: flex;
            align-items: center;
            justify-content: center;
            color: white;
            font-weight: bold;
        }}
        .stats {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
            gap: 15px;
            margin: 20px 0;
        }}
        .stat-card {{
            background: #f8f9fa;
            padding: 15px;
            border-radius: 5px;
            text-align: center;
        }}
        .stat-value {{
            font-size: 24px;
            font-weight: bold;
            color: #3498db;
        }}
        .stat-label {{
            font-size: 12px;
            color: #7f8c8d;
            margin-top: 5px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🚀 Import PPM Tasks to AGL-Tracking</h1>
        
        <div class="info">
            <strong>📋 Ready to Import:</strong> {len(tasks)} PPM tasks from Excel file
            <br><br>
            <strong>📁 Source:</strong> 11445575_with_shifts.xlsx
        </div>
        
        <div class="stats">
            <div class="stat-card">
                <div class="stat-value">{len(tasks)}</div>
                <div class="stat-label">Total Tasks</div>
            </div>
        </div>
        
        <div class="warning">
            <strong>⚠️ Important:</strong>
            <ul>
                <li>This will add {len(tasks)} new PPM tasks to your system</li>
                <li>Make sure Firebase is configured and connected</li>
                <li>Existing tasks will NOT be affected</li>
                <li>You can review tasks after import and delete if needed</li>
            </ul>
        </div>
        
        <div id="progress">
            <div class="progress-bar">
                <div class="progress-fill" id="progressFill">0%</div>
            </div>
            <p id="progressText">Importing tasks...</p>
        </div>
        
        <div class="success" id="successMessage">
            <strong>✅ Success!</strong> All tasks have been imported successfully!
            <br><br>
            <a href="index.html">Go to Main Application →</a>
        </div>
        
        <div id="buttons">
            <button onclick="importTasks()">Import Tasks Now</button>
            <button class="danger" onclick="window.location.href='index.html'">Cancel</button>
        </div>
    </div>

    <script src="firebase-config.js"></script>
    <script>
        // Tasks data embedded from Excel import
        const tasksToImport = {json.dumps(tasks, indent=8)};
        
        async function importTasks() {{
            if (!confirm(`Are you sure you want to import ${{tasksToImport.length}} tasks?`)) {{
                return;
            }}
            
            document.getElementById('buttons').style.display = 'none';
            document.getElementById('progress').style.display = 'block';
            
            try {{
                // Initialize Firebase
                if (typeof firebase === 'undefined') {{
                    throw new Error('Firebase SDK not loaded. Please check firebase-config.js');
                }}
                
                const ppmTasksRef = firebase.database().ref('ppmTasks');
                
                let imported = 0;
                const total = tasksToImport.length;
                
                // Import tasks in batches to avoid overwhelming Firebase
                const batchSize = 10;
                for (let i = 0; i < tasksToImport.length; i += batchSize) {{
                    const batch = tasksToImport.slice(i, i + batchSize);
                    
                    // Import batch
                    const promises = batch.map(task => {{
                        const taskRef = ppmTasksRef.child(task.id.toString());
                        return taskRef.set(task);
                    }});
                    
                    await Promise.all(promises);
                    
                    imported += batch.length;
                    const percent = Math.round((imported / total) * 100);
                    
                    document.getElementById('progressFill').style.width = percent + '%';
                    document.getElementById('progressFill').textContent = percent + '%';
                    document.getElementById('progressText').textContent = 
                        `Imported ${{imported}} of ${{total}} tasks...`;
                    
                    // Small delay between batches
                    await new Promise(resolve => setTimeout(resolve, 100));
                }}
                
                document.getElementById('progress').style.display = 'none';
                document.getElementById('successMessage').style.display = 'block';
                
                console.log(`✅ Successfully imported ${{imported}} tasks!`);
                
            }} catch (error) {{
                console.error('❌ Import error:', error);
                alert('Import failed: ' + error.message);
                document.getElementById('buttons').style.display = 'block';
                document.getElementById('progress').style.display = 'none';
            }}
        }}
    </script>
</body>
</html>
"""
    
    with open('import_tasks.html', 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"\n📄 Import interface created: import_tasks.html")
    print(f"   Open this file in your browser to import tasks into AGL-Tracking")

if __name__ == "__main__":
    # Import tasks from Excel
    tasks = import_tasks_from_excel("11445575_with_shifts.xlsx")
    
    # Export to JSON
    export_to_json(tasks)
    
    # Generate HTML import interface
    generate_import_html(tasks)
    
    print(f"\n{'=' * 80}")
    print(f"🎯 Next Steps:")
    print(f"{'=' * 80}")
    print(f"1. Review the imported tasks in 'imported_ppm_tasks.json'")
    print(f"2. Open 'import_tasks.html' in your browser")
    print(f"3. Click 'Import Tasks Now' to add them to AGL-Tracking")
    print(f"4. Go to main application (index.html) to view imported tasks")
    print(f"{'=' * 80}")
