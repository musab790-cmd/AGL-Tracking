#!/usr/bin/env python3
"""Parse Excel file to understand its structure and extract task information."""

import openpyxl
import json
from datetime import datetime

def parse_excel(filename):
    """Parse the Excel file and extract task information."""
    wb = openpyxl.load_workbook(filename)
    
    print(f"📊 Excel File Analysis: {filename}")
    print(f"=" * 80)
    print(f"\nSheet Names: {wb.sheetnames}")
    
    results = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'=' * 80}")
        print(f"📄 Sheet: {sheet_name}")
        print(f"{'=' * 80}")
        
        # Get dimensions
        max_row = ws.max_row
        max_col = ws.max_column
        print(f"Dimensions: {max_row} rows x {max_col} columns")
        
        # Read headers (first row)
        headers = []
        for col in range(1, max_col + 1):
            cell_value = ws.cell(1, col).value
            headers.append(cell_value)
        
        print(f"\nHeaders ({len(headers)} columns):")
        for idx, header in enumerate(headers, 1):
            print(f"  {idx}. {header}")
        
        # Read all data rows
        data_rows = []
        for row in range(2, max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                cell_value = ws.cell(row, col_idx).value
                if header:  # Only add if header exists
                    row_data[header] = cell_value
            data_rows.append(row_data)
        
        print(f"\nTotal Data Rows: {len(data_rows)}")
        
        # Show first few rows as sample
        print(f"\n📋 Sample Data (first 3 rows):")
        for idx, row in enumerate(data_rows[:3], 1):
            print(f"\n  Row {idx}:")
            for key, value in row.items():
                if value is not None:
                    print(f"    {key}: {value}")
        
        results[sheet_name] = {
            'headers': headers,
            'data': data_rows
        }
    
    # Save to JSON for easy inspection
    output_file = 'parsed_tasks.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        # Convert datetime objects to strings for JSON serialization
        def default_serializer(obj):
            if isinstance(obj, datetime):
                return obj.isoformat()
            return str(obj)
        
        json.dump(results, f, indent=2, default=default_serializer, ensure_ascii=False)
    
    print(f"\n{'=' * 80}")
    print(f"✅ Data exported to: {output_file}")
    print(f"{'=' * 80}")
    
    return results

if __name__ == "__main__":
    parse_excel("11445575_with_shifts.xlsx")
